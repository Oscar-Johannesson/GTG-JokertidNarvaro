from msal import ConfidentialClientApplication
from nicegui import ui
from fastapi import Request
from Backend import log, EvaluateMail, EvaluateRank, getUsername
from backbackend import loadClampedNarvaro, antal_som_narvarande
import random
import time
from datetime import datetime
from openpyxl import load_workbook





#STOP! if you are not a developer this is the line you do not cross -> ----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def checktoken(otoken):
    email = reademailfromoliversAss(otoken)
    if email == "NoFuckYou" or email == None:
        ui.navigate.to('/')
        return False
    else:
        return True


def load_excel():
    file_path = './Data/Narvaro/Narvaro_2024.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []

    for row in sheet.iter_rows(values_only=True):
        if row != None:
            data.append(row)
        else:
            data.append(' ')

    return data


def load_day_excel(load, date = datetime.now().strftime('%Y-%m-%d')):
    if date == None:
        date = datetime.now().strftime('%Y-%m-%d')
    if load == False:
        print('')
    elif load == 'r':
        okay = loadClampedNarvaro(date)
        if okay != 0:
            return 0
        ui.run_javascript('location.reload();')
    else:
        okay = loadClampedNarvaro(date)
        if okay != 0:
            return 0


    file_path = './Data/Narvaro/DagNarvaro.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []

    for row in sheet.iter_rows(values_only=True):
        if row != None:
            data.append(row)
        else:
            data.append(' ')
    
    
    return data



allowInlog = True
allownarvaro = True
@ui.page('/user/{otoken}')
def page(otoken: str):
    if checktoken(otoken) == False:
        main_menu('Access Denied: Nice try')
    ui.timer(5.0, lambda: checktoken(otoken))


  #  ui.query('body').classes('bg-gradient-to-t from-green-400 to-blue-200')

    image1 = 'https://wallpapers-clan.com/wp-content/uploads/2024/10/noble-shiba-contemplation-blue-sky-desktop-wallpaper-cover.jpg'
    image2 = 'https://wallpapers-clan.com/wp-content/uploads/2023/12/dolphin-cow-jumping-sea-meme-desktop-wallpaper-cover.jpg'
    image3 = 'https://wallpapers-clan.com/wp-content/uploads/2024/10/northern-lights-mountain-view-desktop-wallpaper-cover.jpg'
    select = random.randrange(1, 4)
    if select == 1:
        image = image1
    elif select == 2:
        image = image2
    elif select == 3:
        image = image3
    else:
        image = image3
    ui.add_head_html(f'''
        <style>
        body {{
            background-image: url({image});
            background-size: cover;
            background-position: center;
            background-repeat: no-repeat;
        }}
        </style>
        ''')
    ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')

    with ui.element('div').classes('h-screen w-full flex flex-col justify-between'):
        ui.markdown('Online - L\u00E4rarpanel').style('font-size: 30px; text-align: center; margin-top: 20px;')

        with ui.element('div').classes('flex absolute bottom-10 left-0 transform space-x-2'):
                ui.button('logout', on_click=lambda: logoutuser(otoken, None))
                ui.button('Restart Server', on_click=lambda: os.system('shutdown /r /t 0'))
                ui.button('Shutdown Server', on_click=lambda: os.system('shutdown /t 0'))
        
        with ui.element('div').classes('absolute top-32 right-0 w-1/3 h-4/5 overflow-auto'):
            data = load_excel() 
            with ui.element('div').classes('p-2 bg-blue-100 h-full overflow-auto'):
                with ui.scroll_area().classes('h-full border'):
                    for row in data:
                        ui.markdown(' ▒ '.join(map(str, row)))

        with ui.element('div').classes('absolute top-32 right-1/3 w-1/3 h-4/5 overflow-auto'):
            data2 = load_day_excel(False)
            with ui.element('div').classes('p-2 bg-blue-100 h-full overflow-auto'):
                with ui.scroll_area().classes('h-full border'):
                    for row in data2:
                        ui.markdown(' ▒ '.join(map(str, row)))

        with ui.element('div').classes('absolute top-32 right-2/3  overflow-auto'):
            with ui.element('div').classes('p-2 bg-blue-100 h-full overflow-auto'):
                personer = antal_som_narvarande()
                ui.markdown('**Totala personer:**     ')
                ui.markdown(str(personer)).style('text-align: center;')
            today_date = datetime.now().strftime('%Y-%m-%d')
            with ui.button('v\u00E4lj datum', on_click=lambda: menu.open()).classes('cursor-pointer') as button:
                with ui.menu().props('no-parent-event') as menu:
                    with ui.date(today_date, on_change=lambda e: load_day_excel('r', e.value)):
                        with ui.row().classes('justify-end'):
                            ui.button('Close', on_click=menu.close).props('flat')


def logoutuser(otoken=None, email=None):

    ui.navigate.to('/')

    filename = "./Data/SessionTokens.txt"
    
    with open(filename, 'r') as f:
        tokens = f.readlines()

    updated_tokens = []
    found = False
    
    for row in tokens:
        if (otoken and otoken in row) or (email and email in row):
            found = True 
            continue  
        updated_tokens.append(row)

    if found:
        with open(filename, 'w') as f:
            f.writelines(updated_tokens)



def store_username(username):
    ui.run_javascript(f"localStorage.setItem('username', '{username}');")
    ui.notify(f'Username "{username}" has been stored!')

def retrieve_username():
    ui.run_javascript("""
        const username = localStorage.getItem('username');
    """)


CLIENT_ID = "a9e7a7d7-f1c7-42f7-b2a1-940a213baa07"
TENANT_ID = "4f4427c9-47c4-4f08-bbf2-4ddf75a64b3b"
CLIENT_SECRET = "fu88Q~GAkm.jq3OnGj25YYR.zLFuCgbWIPWMEcH1"
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
REDIRECT_URI = "https://PW0BAGDM/login-redirect" # http://localhost:8888/login-redirect
SCOPE = ["User.Read"]

app = ConfidentialClientApplication(
    CLIENT_ID,
    authority=AUTHORITY,
    client_credential=CLIENT_SECRET
)


def readtocken(email):
    filename = "./Data/SessionTokens.txt"
    
    with open(filename, 'r') as f:
        tokens = f.readlines()

    matching_row = None
    
    for row in tokens:
        if email in row:
            matching_row = row
            break
    
    if matching_row:
        token = matching_row.split(',')[1].strip()
        return token
    else:
        return "NOFuckYou"


def reademailfromoliversAss(token):
    filename = "./Data/SessionTokens.txt"
    
    with open(filename, 'r') as f:
        tokens = f.readlines()

    matching_row = None
    
    for row in tokens:
        if token in row:
            matching_row = row
            break
    
    if matching_row:
        email = matching_row.split(',')[0].strip()
        return email
    else:
        return None




def CheckEmail(emailentered: str):

    mailaddresvalue = EvaluateMail(emailentered)
    if mailaddresvalue == 1 or mailaddresvalue == 0:
        log(f"stopped invalid attempt, {emailentered}")
        return False
    
    try:
        file = f'./Data/{["Deltagare", "Larare", "Admin"][mailaddresvalue - 1]}.txt'
    except:
        log(f"online: did not rescive email data, login blocked, email {emailentered}")
        return False
    
    
    with open(file, 'r') as ids_file:
        ids_lines = ids_file.readlines()

        for row, id_value in enumerate(ids_lines, start=1):
            actual_id, name, email = map(str.strip, id_value.split(' || '))
            if email == emailentered:
                break
            
        if(email != emailentered):
            return False

    try:
        genoratedToken = f"{actual_id}{getUsername(actual_id).split(' ')[0]}{random.randrange(834, 620057)}{random.randrange(742, 310052)}{random.randrange(99, 999999)}-{random.randrange(849, 454600)}{random.randrange(738, 627006)}{random.randrange(876, 940079)}"

        filename = "./Data/SessionTokens.txt"
        tokens = []
        with open(filename, 'r') as f:
            tokens = f.readlines()

        tokens = [token for token in tokens if not token.startswith(email + ',')]
        
        tokens.append(f"{email},{genoratedToken}\n")
        
        with open(filename, 'w') as f:
            f.writelines(tokens)
            return True
    except:
        return False

def start_login():
    auth_url = app.get_authorization_request_url(SCOPE, redirect_uri=REDIRECT_URI)
    ui.run_javascript(f'window.open("{auth_url}", "_self")') 

@ui.page('/login-redirect')
async def login_redirect(request: Request):
    global errorsussy
    auth_code = request.query_params.get('code')
    
    if auth_code:
        token_response = app.acquire_token_by_authorization_code(auth_code, scopes=SCOPE, redirect_uri=REDIRECT_URI)
        if "id_token" in token_response:
            user_info = token_response['id_token_claims']
            user_email = user_info.get('email') or user_info.get('preferred_username')
            
            if CheckEmail(user_email):
                ui.label("redirecting...")
                otoken = readtocken(user_email)
                ui.navigate.to(f'/user/{otoken}')

            else:
                main_menu('Access Denied: You not authorized.')

        else:
            ui.navigate.to('/')
    else:
        ui.notify("Authorization code missing.")


@ui.page('/')
def main_menu(errorsussy = ''):
    ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
    ui.add_head_html('''
    <style>
    body {
        background-image: url('https://assets.volvo.com/is/image/VolvoInformationTechnologyAB/about-us-quality-grey-sky-truck?qlt=82&wid=768&ts=1665660240235&dpr=off&fit=constrain');
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
    }
    </style>
    ''')
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            with ui.element('div').classes('p-2 bg-blue-100'):
                ui.markdown(f'online - l\u00E4rarpanelen').style('font-size: 23.333px; text-align: center;')
                ui.markdown(errorsussy).style('font-size: 13.333px; text-align: center;').tailwind('drop-shadow', 'font-bold', 'text-red-600')
            with ui.button(on_click=start_login, color='white').style('background-color: #0078D4; border: none; border-radius: 4px; display: flex; align-items: center; padding: 10px 15px; cursor: pointer;'):
                ui.image('https://upload.wikimedia.org/wikipedia/commons/4/44/Microsoft_logo.svg').style('width: 24px; height: 24px; margin-right: 10px;')
                ui.label('Logga in med Microsoft').style('font-size: 16px; font-family: Roboto, Arial, sans-serif; color: black; letter-spacing: .2px; line-height: 24px;')








iconImage='https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcS9dtZ1_nUeTqeLf0UzogbGAGVUdmMtkVT9mA&s'

#ui.run(title='GTG - Jokertid närvaro', port=8888, favicon=iconImage)
load_day_excel(True)
#, ssl_cert='./certificate/certificate.crt', ssl_key='./certificate/private.key'
ui.run(
    port=443,
    ssl_certfile="./certificate/certificate.crt",
    ssl_keyfile="./certificate/private.key",
    title='GTG - Jokertid närvaro',
    favicon=iconImage,
    show = False
) 
