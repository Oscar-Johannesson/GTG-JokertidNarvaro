from datetime import datetime
import re
import os
from nicegui.ui import card
import openpyxl
import ctypes
import time
from openpyxl import load_workbook, Workbook
from Backend import log
from Backend import datafile
global datafile


#STOP! if you are not a developer this is the line you do not cross -> ----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------










def loadClampedNarvaro(datevalue): # skapar en fill i "./Data/Narvaro/DagNarvaro.xlsx" som samanfattar det datum du mattar in FELKÅDER: 0= allt bra,  1 = kan inte öppna filen för att skriva till den, 2 = kan inte hitta datumet
    global datafile
    datestring = f"**Närvaro för {str(datevalue)}"

    wb = load_workbook(datafile)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell = sheet.cell(1,1)

    wb2 = Workbook()
    new_sheet = wb2.active

    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row, 1)
        value = cell.value[:24] if isinstance(cell.value, str) else cell.value
        if value == datestring:
            break
        elif row == sheet.max_row:
            new_sheet.cell(1, 1).value = "Saknas data... inte mycket att se"
            new_sheet.cell(2, 1).value = "Oscar was here 🤪"
            try:
                wb2.save('./Data/Narvaro/DagNarvaro.xlsx')
                log(f"Data saved to './Data/Narvaro/DagNarvaro.xlsx' but there was no data...")
                return 0
            except PermissionError as e:
                log(f"Error: {e} cant open the file")
                return 1
            except Exception as e:
                print(f"An unexpected error occurred: {e}")
                return 3
            return 2


    for row2 in range(row + 1, sheet.max_row + 1):
        cell2 = sheet.cell(row2, 1)
        value2 = cell2.value[:13] if isinstance(cell2.value, str) else cell2.value
        if value2 == "**Närvaro för":
            row2 = row2 - 1
            break
        elif row2 == sheet.max_row + 1:
            break

    new_row_idx = 1  
    for readrow in range(row, row2 + 1):
        for col in range(1, sheet.max_column + 1):  
            cell3 = sheet.cell(readrow, col)
            new_sheet.cell(new_row_idx, col).value = cell3.value  
        new_row_idx += 1 

    try:
        wb2.save('./Data/Narvaro/DagNarvaro.xlsx')
        log(f"Data saved to './Data/Narvaro/DagNarvaro.xlsx' ready to read!")
        return 0
    except PermissionError as e:
        log(f"Error: {e} cant open the file")
        return 1
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return 3




def antal_som_narvarande():
    datestring = "**Närvaro för"
    datestring2 = "Tid In (HH/"

    wb = load_workbook('./Data/Narvaro/DagNarvaro.xlsx')
    sheet = wb['Sheet']

    wb2 = Workbook()
    new_sheet = wb2.active

    total = 0
    seen_names = set() 
    for row in range(1, sheet.max_row + 1):
        cell_value_col1 = sheet.cell(row, 1).value
        value_col1 = cell_value_col1[:11] if isinstance(cell_value_col1, str) else cell_value_col1

        cell_value_col2 = sheet.cell(row, 2).value

        if value_col1 != datestring and datestring2 != value_col1 and value_col1 is not None:
            if cell_value_col2 not in seen_names:  
                seen_names.add(cell_value_col2)  
                total += 1
        if value_col1 == "Saknas data":
            return 0

    return total - 1