from nicegui import ui
from openpyxl import load_workbook, Workbook
from Backend import EvaluateRank, fixSvenska, addMember, log
from backbackend import loadClampedNarvaro, antal_som_narvarande
from Backend import getPassword
from Backend import getEmail
from Backend import getUsername
from datetime import datetime
from Backend import fixcardid
from Backend import narvaro
from Backend import error_sound
from Backend import print_current_time
import time
from Backend import LookupNarvaro, press_backspace
import asyncio
import os
from Backend import datafile
global datafile

#STOP! if you are not a developer this is the line you do not cross -> ----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------










nykontocardID = "0"
nykontoNamn = "0"
nykontoMail = "0"
nykontoPass = "0"
nykontoRank = 0
nykontoAuth = "0"

allowInlog = True
allownarvaro = True



@ui.page('/larare3')
def sida8():
    time.sleep(1)
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.markdown('konfirmera med admin kort').style('font-size: 30px; text-align: center;')
            ui.markdown('f\xf6r att skapa l\u00E4rar konto').style('font-size: 20px; text-align: center;')
            cardbox = ui.input('', validation={'Too long': lambda value: len(value) <= 9}, on_change=lambda e: skapare(e.value, 2)).props('type=password')
            ui.timer(0.1, focus_input)


@ui.page('/admin3')
def sida7():
    time.sleep(1)
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.markdown('konfirmera med admin kort').style('font-size: 30px; text-align: center;')
            ui.markdown('f\xf6r att skapa admin konto').style('font-size: 20px; text-align: center;')
            cardbox = ui.input('', validation={'Too long': lambda value: len(value) <= 9}, on_change=lambda e: skapare(e.value, 3)).props('type=password')
            ui.timer(0.1, focus_input)


@ui.page('/larare2')
def larare2():
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.markdown('Skanna ditt kort').style('font-size: 30px; text-align: center;')
            cardbox = ui.input('', validation={'Too long': lambda value: len(value) <= 9}, on_change=lambda e: larar_skapare(e.value, 2)).props('type=password')
            ui.timer(0.1, focus_input)


@ui.page('/admin2')
def admin2():
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.markdown('Skanna ditt kort').style('font-size: 30px; text-align: center;')
            cardbox = ui.input('', validation={'Too long': lambda value: len(value) <= 9}, on_change=lambda e: admin_skapare(e.value, 3)).props('type=password')
            ui.timer(0.1, focus_input)


@ui.page('/deltagare2')
def deltagare2():
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.markdown('Skanna ditt kort').style('font-size: 30px; text-align: center;')
            cardbox = ui.input('', validation={'Too long': lambda value: len(value) <= 9}, on_change=lambda e: Deltag_skapare(e.value, 1)).props('type=password')
            ui.timer(0.1, focus_input)



@ui.page('/larare')
def larare():
    global nykontoNamn
    global nykontoMail
    global nykontoPass
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.input("Namn och efternamn", on_change=lambda e: handle_changeNamn(e)).props('')
            ui.input("email", on_change=lambda e: handle_changeMail(e)).props('model="email" type="email"')
            ui.input('password', on_change=lambda e: handle_changePass(e)).props('type=password')
            ui.button('skapa konto', on_click=lambda: check_Larare())


@ui.page('/admin')
def admin():
    global nykontoNamn
    global nykontoMail
    global nykontoPass
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.input("Namn och efternamn", on_change=lambda e: handle_changeNamn(e)).props('')
            ui.input("email", on_change=lambda e: handle_changeMail(e)).props('model="email" type="email"')
            ui.input('password', on_change=lambda e: handle_changePass(e)).props('type=password')
            ui.button('skapa konto', on_click=lambda: check_Admin())


@ui.page('/deltagare')
def deltagare():
    global nykontoNamn
    global nykontoMail
    global nykontoPass
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.input("Namn och efternamn", on_change=lambda e: handle_changeNamn(e)).props('')
            ui.input("email", on_change=lambda e: handle_changeMail(e)).props('model="email" type="email"')
            ui.input('password', on_change=lambda e: handle_changePass(e)).props('type=password')
            ui.button('skapa konto', on_click=lambda: check_deltagare())
            
def handle_changeNamn(e):
    global nykontoNamn
    nykontoNamn = e.value

def handle_changeMail(e):
    global nykontoMail
    nykontoMail = e.value

def handle_changePass(e):
    global nykontoPass
    nykontoPass = e.value
    
def check_deltagare():
    global nykontoNamn
    global nykontoMail
    global nykontoPass
    #check mail och om alla har data
    if nykontoNamn != "0" and nykontoMail != "0" and nykontoPass != "0":
        #chech email here 
        ui.navigate.to('/deltagare2')
    else:
        ui.notify('Please fill out all of the infomation!')

def check_Larare():
    global nykontoNamn
    global nykontoMail
    global nykontoPass
    #check mail och om alla har data
    if nykontoNamn != "0" and nykontoMail != "0" and nykontoPass != "0":
        #chech email here 
        ui.navigate.to('/larare2')
    else:
        ui.notify('Please fill out all of the infomation!')

def check_Admin():
    global nykontoNamn
    global nykontoMail
    global nykontoPass
    #check mail och om alla har data
    if nykontoNamn != "0" and nykontoMail != "0" and nykontoPass != "0":
        #chech email here 
        ui.navigate.to('/admin2')
    else:
        ui.notify('Please fill out all of the infomation!')

@ui.page('/kontot_skapat')
def kontot_skapat():
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.markdown('konto skapat').style('font-size: 30px; text-align: center;')
            time.sleep(3)
            ui.navigate.to('/loading')

@ui.page('/konto_errore')
def konto_errore():
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.markdown('Kunde inte skapa konto').style('font-size: 30px; text-align: center;')
            ui.markdown('oscar was here').style('font-size: 15px; text-align: center;')
            time.sleep(3)
            ui.navigate.to('/loading')


def Deltag_skapare(value, rank):
    if len(value) >= 9:
        global nykontocardID
        global nykontoRank
        global nykontoNamn
        global nykontoMail
        global nykontoPass
        global nykontoAuth

        nykontoAuth = "0"
        nykontoRank = rank
        nykontocardID = value[:9]
        
        print(str(nykontocardID), str(nykontoNamn), str(nykontoMail), int(nykontoRank), str(nykontoPass), str(nykontoAuth))
        try:
            result = addMember(str(nykontocardID), str(nykontoNamn), str(nykontoMail), int(nykontoRank), str(nykontoPass), str(nykontoAuth))
            if result == 0 or result == "Label":
                ui.navigate.to('/konto_errore')
            else:
                ui.navigate.to('/kontot_skapat')
        except:
            ui.navigate.to('/errore')


def skapare(value, rank):
    if len(value) >= 9:
        global nykontocardID
        global nykontoRank
        global nykontoNamn
        global nykontoMail
        global nykontoPass
        global nykontoAuth

        nykontoRank = rank
        nykontoAuth = value[:9]
        
        print(str(nykontocardID), str(nykontoNamn), str(nykontoMail), int(nykontoRank), str(nykontoPass), str(nykontoAuth))
        try:
            result = addMember(str(nykontocardID), str(nykontoNamn), str(nykontoMail), int(nykontoRank), str(nykontoPass), str(nykontoAuth))
            print(result)
            if result == 0 or result == "Label":
                ui.navigate.to('/konto_errore')
            else:
                ui.navigate.to('/kontot_skapat')
        except:
            ui.navigate.to('/errore')

            
def larar_skapare(value, rank):
    if len(value) >= 9:
        cardID = value
        global nykontocardID
        nykontocardID = cardID[:9]
        ui.navigate.to('/loadingAUTHLARARE')
       # time.sleep(2)
       # ui.navigate.to('/larare3')

def admin_skapare(value, rank):
    if len(value) >= 9:
        cardID = value
        global nykontocardID
        nykontocardID = cardID[:9]
        ui.navigate.to('/loadingAUTHARMIN')
       # time.sleep(2)
       # ui.navigate.to('/admin3')

@ui.page('/loadingAUTHARMIN')
def loadingAUTH():
    ui.query('body').classes('bg-gray-400')
    with ui.column().style('height: 100vh; display: flex; align-items: center; justify-content: center; width: 100%;'):
        ui.spinner(size='150px')
    ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
    ui.navigate.to('/admin3')

@ui.page('/loadingAUTHLARARE')
def loadingAUTH():
    ui.query('body').classes('bg-gray-400')
    with ui.column().style('height: 100vh; display: flex; align-items: center; justify-content: center; width: 100%;'):
        ui.spinner(size='150px')
    ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
    ui.navigate.to('/larare3')



@ui.page('/konto_skapare')
def konto_skapare():
        with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
            with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
                ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
                ui.query('body').classes('bg-gradient-to-t from-cyan-400 to-cyan-100')
                ui.markdown('vill du skapa ett konto?').style('font-size: 34.333px; text-align: center;')
                with ui.button_group():
                    ui.button('deltagar skapare', on_click=lambda: ui.navigate.to('/deltagare'))
                    ui.button('l\u00E4rar skapare', on_click=lambda: ui.navigate.to('/larare'))
                    ui.button('admin skapare', on_click=lambda: ui.navigate.to('/admin'))
                ui.button('back', on_click=ui.navigate.back)





@ui.page('/larar_inlogg')
def larar_inlogg():
    global cardsboxs
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.markdown('skanna l\u00E4rar kort').style('font-size: 30px; text-align: center;')
            ui.markdown('Endast L\u00E4rare').style('font-size: 20px; text-align: center;')
            cardsboxs = ui.input('', validation={'Too long': lambda value: len(value) <= 9}, on_change=lambda e: lararlosen(e.value)).props('type=password')
            ui.button('back', on_click=lambda: ui.navigate.to('/'))
            ui.timer(0.1, focus_input)



def lararlosen(value):
    global cardsboxs
    if len(value) >=9:
        cardsboxs.set_value('')
        cardsboxs.set_value('')
        cardID=value[:9]
        if EvaluateRank(cardID) >= 2:
            log(f"Användare {getUsername(cardID)} has logged in to lararpanel, cardID: {cardID}")
            ui.navigate.to('/lararpanel')
            cardsboxs.set_value('')
        else:
            log (f"user {getUsername(cardID)} entered a non admin cardID {cardID}, going back to root")
            time.sleep(1)
            ui.notify('nuh uh u no l\u00E4rare')
            cardsboxs.set_value('')


@ui.page('/deltagar_inlogg')
def deltagar_inlogg():
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
            with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
                ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
                ui.query('body').classes('bg-gray-400')
                ui.markdown(f'**Delragar inlogg**').style('font-size: 34.333px; text-align: center;')
                ui.markdown(f'Kommer snart....').style('font-size: 23.333px; text-align: center;')
                ui.button('Back', on_click=lambda: ui.navigate.to('/'))


cardID = "0"


@ui.page('/sparar_narvaro')
async def sparar_narvaro():
    global cardID
    namn = "namn"
    name = fixSvenska(getUsername(cardID))
    inputbox.set_value('')
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.query('body').classes('bg-green-400')
            ui.markdown(f'V\u00E4lkommen {name}').style('font-size: 34.333px; text-align: center;')
            ui.markdown(f'N\u00e4rvaro sparad klockan {datetime.now():%X}').style('font-size: 23.333px; text-align: center;')
            time.sleep(0.7)
            ui.navigate.to('/loading')

@ui.page('/errore/{code}')
def errore(code: str):
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.query('body').classes('bg-red')
            ui.markdown(f"Error {code}").style('font-size: 34.333px; text-align: center;')
            ui.markdown('Ett fel har uppst\u00E5tt').style('font-size: 34.333px; text-align: center;')
            error_sound()
            processing = False


    
@ui.page('/avgang')
async def avgang():
    inputbox.set_value('')
    global cardID
    name = "namn"
    name = fixSvenska(getUsername(cardID))
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.query('body').classes('bg-orange-400')
            ui.markdown(f'Hej d\u00E5 {name}!').style('font-size: 34.333px; text-align: center;')
            ui.markdown(f'N\u00e4rvaro sparad klockan {datetime.now():%X}').style('font-size: 23.333px; text-align: center;')
            time.sleep(0.7)
            ui.navigate.to('/loading') 

@ui.page('/loading')
def loading():
    global processing
    ui.query('body').classes('bg-gray-400')
    ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
    with ui.column().style('height: 100vh; display: flex; align-items: center; justify-content: center; width: 100%;'):
        ui.spinner(size='150px')
        time.sleep(1)
        processing = False
    ui.navigate.to('/') 
    
@ui.page('/loading_two')
def loading_two():
    ui.query('body').classes('bg-gray-400')
    with ui.column().style('height: 100vh; display: flex; align-items: center; justify-content: center; width: 100%;'):
        ui.spinner(size='150px')
    ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
    ui.navigate.to('/sparar_narvaro')


@ui.page('/loading_three')
def loading_three():
    ui.query('body').classes('bg-gray-400')
    with ui.column().style('height: 100vh; display: flex; align-items: center; justify-content: center; width: 100%;'):
        ui.spinner(size='150px')
    ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
    ui.navigate.to('/avgang')


def load_excel():
    global datafile
    file_path = datafile
    #file_path = './Data/Narvaro/Narvaro_2024.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []

    for row in sheet.iter_rows(values_only=True):
        if row != None:
            data.append(row)
        else:
            data.append(' ')

    return data


def load_day_excel(load, date = datetime.now().strftime('%Y-%m-%d')):
    if date == None:
        date = datetime.now().strftime('%Y-%m-%d')
    if load == False:
        print('')
    elif load == 'r':
        okay = loadClampedNarvaro(date)
        if okay != 0:
            return 0
        ui.run_javascript('location.reload();')
    else:
        okay = loadClampedNarvaro(date)
        if okay != 0:
            return 0


    file_path = './Data/Narvaro/DagNarvaro.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []

    for row in sheet.iter_rows(values_only=True):
        if row != None:
            data.append(row)
        else:
            data.append(' ')
    
    
    return data

@ui.page('/lararpanel')
def lararpanel():
    log(f"lararpanel opend...")
    ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
    ui.query('body').classes('bg-gradient-to-t from-green-400 to-blue-200')

    with ui.element('div').classes('h-screen w-full flex flex-col justify-between'):
        ui.markdown('L\u00E4rarpanel').style('font-size: 30px; text-align: center; margin-top: 20px;')

        with ui.element('div').classes('flex absolute bottom-10 left-0 transform space-x-2'):
                ui.button('Back', on_click=lambda: ui.navigate.to('/'))
        
        with ui.element('div').classes('absolute top-32 right-0 w-1/3 h-4/5 overflow-auto'):
            data = load_excel() 
            with ui.element('div').classes('p-2 bg-blue-100 h-full overflow-auto'):
                with ui.scroll_area().classes('h-full border'):
                    for row in data:
                        ui.markdown(' ▒ '.join(map(str, row)))

        with ui.element('div').classes('absolute top-32 right-1/3 w-1/3 h-4/5 overflow-auto'):
            data2 = load_day_excel(False)
            with ui.element('div').classes('p-2 bg-blue-100 h-full overflow-auto'):
                with ui.scroll_area().classes('h-full border'):
                    for row in data2:
                        ui.markdown(' ▒ '.join(map(str, row)))

        with ui.element('div').classes('absolute top-32 right-2/3  overflow-auto'):
            with ui.element('div').classes('p-2 bg-blue-100 h-full overflow-auto'):
                personer = antal_som_narvarande()
                ui.markdown('**Totala personer:**     ')
                ui.markdown(str(personer)).style('text-align: center;')
            today_date = datetime.now().strftime('%Y-%m-%d')
            with ui.button('v\u00E4lj datum', on_click=lambda: menu.open()).classes('cursor-pointer') as button:
                with ui.menu().props('no-parent-event') as menu:
                    with ui.date(today_date, on_change=lambda e: load_day_excel('r', e.value)):
                        with ui.row().classes('justify-end'):
                            ui.button('Close', on_click=menu.close).props('flat')


@ui.page('/adminpanel')
def adminpanel():
    log(f"adminpannel opend...")
    global allownarvaro
    with open('./Data/Logs/latestLog.txt', 'r') as loggfil:
        filenlogg = loggfil.read()
    ui.query('body').classes('bg-gradient-to-t from-blue-400 to-blue-200')

    ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')

    with ui.element('div').classes('h-screen w-full flex flex-col justify-between'):
        ui.markdown('Adminpanel').style('font-size: 40px; text-align: center; margin-top: 20px;')
        
        with ui.element('div').classes('absolute top-16 left-4 flex items-center space-x-4'):
            switch1 = ui.switch('stoppa inlogg', value=not allowInlog, on_change=sawawawawawpppa)
            kapp = ui.switch('stoppa inst\u00E4mpling', value=not allownarvaro, on_change=sapiswapilillaoliver)
        
        with ui.element('div').classes('flex absolute bottom-10 left-0 transform space-x-2'):
            ui.button('restart', on_click=lambda: os.system('shutdown /r /t 0'))
            ui.button('shutdown', on_click=lambda: os.system('shutdown /t 0'))

        with ui.element('div').classes('flex absolute bottom-10 left-1/2 transform -translate-x-1/2 space-x-2'):
            ui.button('back', on_click=lambda: ui.navigate.to('/'))

        with ui.element('div').classes('absolute top-32 right-0 w-1/2 h-5/6 overflow-auto'):
                with ui.scroll_area().classes('h-5/6 border'):
                    ui.markdown(filenlogg.replace('\n', '  \n'))

global cardsboxs

@ui.page('/adminpanellosen')
def adminpanellosen():
    global cardsboxs
    with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
        with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
            ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
            ui.markdown('skanna admin kort').style('font-size: 30px; text-align: center;')
            ui.markdown('gor det nooo').style('font-size: 20px; text-align: center;')
            cardsboxs = ui.input('', validation={'Too long': lambda value: len(value) <= 9}, on_change=lambda e: adminlosen(e.value)).props('type=password')
            ui.button('back', on_click=lambda: ui.navigate.to('/'))
            ui.timer(0.1, focus_input)



def adminlosen(value):
    global cardsboxs
    if len(value) >=9:
        cardsboxs.set_value('')
        cardsboxs.set_value('')
        cardID=value[:9]
        if EvaluateRank(cardID) == 3:
            log(f"Användare {getUsername(cardID)} has logged in to Adminpanel, cardID: {cardID}")
            ui.navigate.to('/adminpanel')
            cardsboxs.set_value('')
        else:
            log (f"user {getUsername(cardID)} entered a non admin cardID {cardID}, going back to root")
            time.sleep(1)
            ui.notify('nuh uh u no have admin')
            cardsboxs.set_value('')



def sapiswapilillaoliver():
    global allownarvaro
    if allownarvaro == True: allownarvaro = False
    else: allownarvaro = True 

def sawawawawawpppa():
    global allowInlog
    if allowInlog == True: allowInlog = False
    else: allowInlog = True 

@ui.page('/nykonto')
def nykonto():
    global processing
    processing = False
    ui.run_javascript("setTimeout(() => { window.location.href = '/'; }, 10000);")
    ui.query('body').classes('bg-gradient-to-t from-blue-400 to-blue-200')
    ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
    with ui.column().style('height: 100vh; display: flex; align-items: center; justify-content: center; width: 100%;'):
       ui.markdown('hittade inte ett kontot, vändligen skapa ett nytt deltagarkonto').style('font-size: 34.333px; text-align: center;')
       ui.button('back', on_click=ui.navigate.back)
       ui.button('skapa konto', on_click=lambda: ui.navigate.to(konto_skapare))


def clearinput():
    a = 0
    while (a > 3):
        inputbox.set_value('')
        time.sleep(0.1)

processing = False

def readcard(value):
    global allownarvaro
    global cardID
    global processing
    if len(value) >= 9:
        k = 0  
        if allownarvaro == False:
            ui.notify("inloggning \u00E4r tillf\u00E4light begr\u00E4nsat")
            log(f"User {getUsername(cardID)} tried to log in but system is locked, cardID: {cardID}")
        else:
            inputbox.set_value('')
            if processing == False:
                inputbox.set_value('')
                cardID = value[:9]
                processing = True
                ui.navigate.to(loading)
                narvaroo = LookupNarvaro(cardID)
                if(narvaroo == 1):
                    ui.navigate.to(loading_two)
                elif(narvaroo == 0):
                    ui.navigate.to(loading_three)
                elif(narvaroo == 2): #för att skapa ett nytt konto
                    processing = False
                    ui.navigate.to(nykonto)
                elif(narvaroo == 4):
                    ui.navigate.to('/errore/1001')
                else:
                    ui.navigate.to(errore) 
    inputbox.set_value('')




with ui.row().classes('w-full items-center'):
    result = ui.label().classes('mr-auto')
    with ui.button(icon='menu'):
        with ui.menu() as menu:
            ui.menu_item('skapa konto', on_click=lambda: ui.navigate.to(konto_skapare))
            ui.menu_item('deltagar inlogg', on_click=lambda: ui.navigate.to(deltagar_inlogg))
            ui.menu_item('l\u00E4rarpanel', on_click=lambda: ui.navigate.to(larar_inlogg))
            ui.menu_item('adminpanel', on_click=lambda: ui.navigate.to(adminpanellosen))
ui.query('body').classes('bg-gradient-to-t from-blue-400 to-blue-200')
with ui.element('div').classes('h-screen w-full flex justify-center items-center overflow-hidden'):
    with ui.column().style('height: 100vh; align-items: center; justify-content: center;'):
        ui.markdown('**GTG** n\u00E4rvaro Jokertid').style('font-size: 34.333px; text-align: center;')
        ui.markdown('Skanna ditt kort').style('font-size: 30px; text-align: center;')
        inputbox = ui.input('', validation={'Too long': lambda value: len(value) <= 9}, on_change=lambda e: readcard(e.value)).props('type=password')
ui.add_head_html('<style>html, body { height: 100%; margin: 0; overflow: hidden; }</style>')
label = ui.label().style('position: fixed; left: 0; bottom: 0; width: 100%; text-align: center; padding: 20px')
ui.label('©2024 Oscar och Oliver: T1s').style('position: fixed; left: 100; bottom: 0; width: 100%; text-align: bottom; padding: 20px')
ui.timer(1.0, lambda: label.set_text(f'{datetime.now():%X}'))

def focus_input():
    ui.run_javascript('document.querySelector("input").focus();')
    ui.run_javascript('document.querySelector("cardbox").focus();')

ui.timer(0.1, focus_input)


ui.run(title='Jokertid n\u00E4rvaro', show = False, port=8080)
