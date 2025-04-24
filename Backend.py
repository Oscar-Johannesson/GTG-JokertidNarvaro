from datetime import datetime
import re
import os
from nicegui.ui import card
import openpyxl
import ctypes
import time
from openpyxl import load_workbook
import pygame




#MAINTENANCE OVER HERE!!!
Filename = 'Narvaro_2024.xlsx'

global datafile
datafile = f'./Data/Narvaro/{Filename}'




#STOP! if you are not a developer this is the line you do not cross -> ----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

def error_sound():
    pygame.mixer.init()
    pygame.mixer.music.load('error.mp3')
    pygame.mixer.music.play()
    
    while pygame.mixer.music.get_busy(): 
        time.sleep(1)

KEYEVENTF_KEYDOWN = 0x0000
KEYEVENTF_KEYUP = 0x0002
VK_BACK = 0x08  

def press_backspace():
    ctypes.windll.user32.keybd_event(VK_BACK, 0, KEYEVENTF_KEYDOWN, 0)  # Press backspace
    ctypes.windll.user32.keybd_event(VK_BACK, 0, KEYEVENTF_KEYUP, 0)    # Release backspace

## oliver detta behöver du:
## from Backend import EvaluateRank
## from Backend import getPassword
## from Backend import getEmail
## from Backend import getUsername

"""
instalera för att det ska funka ;)
pip install openpyxl
pip3 install nicegui
pip install pygame


"""





def changepassword(cardID, newPass):
    found_row = None

    try:
        rank = EvaluateRank(cardID)
        if rank == 0:
            log(f"Can't find that cardID when changing password, cardID: {cardID}")
            return 0
    except Exception as e:
        log(f"Could not find rank when changing password, cardID: {cardID}, error: {str(e)}")
        return 0

    file = f'./Data/{["Deltagare", "Larare", "Admin"][rank - 1]}.txt'
    passfile = file + ':passwords.txt'

    try:
        with open(file, 'r') as sf:
            lines = sf.readlines()
            for idx, line in enumerate(lines):
                if str(cardID) in line:
                    found_row = idx
                    break

        if found_row is not None:
            with open(passfile, 'r') as pf:
                pass_lines = pf.readlines()

            pass_lines[found_row] = str(newPass) + '\n'

            with open(passfile, 'w') as of:
                of.writelines(pass_lines)

            log(f"Password updated successfully for cardID: {cardID}")
            return 1
        else:
            log(f"CardID '{cardID}' not found in the main file.")
            return 0
    except Exception as e:
        log(f"Error while changing password for cardID: {cardID}, error: {str(e)}")
        return 0







def fixSvenska(word):
    try:
        word = word.encode('latin1').decode('utf-8')
    except UnicodeError:
        pass

    swedish_char_map = {
        'å': 'å',
        'ä': 'ä',
        'ö': 'ö',
        'Å': 'Å',
        'Ä': 'Ä',
        'Ö': 'Ö'
    }

    return ''.join([swedish_char_map.get(char, char) for char in word])


cardIDs = 0
def fixcardid(value):
    if len(value) >= 9:
        k = 0
        return value 
        while k < len(value):
            k += 1
            press_backspace()   
            

        

def print_current_time(): #gives the current time for other programs (like log)

    now = datetime.now()
    
    formatted_time = now.strftime("%Y-%m-%d %H:%M:%S")
    
    return(formatted_time)

def filecheck(file, cardID):
    with open(file, "r") as deltagareIDN:
        for line_number, line in enumerate(deltagareIDN):
            maches = re.finditer(cardID, line)
            for match in maches:
                return True



def getUsername(cardID): #hämtar användar namn för att kuna "become the stalker"
    rank = EvaluateRank(cardID)
    if rank == 0:
        log("Invalid rank for the cardID")

    file = f'./Data/{["Deltagare", "Larare", "Admin"][rank - 1]}.txt'

    try:
        with open(file, 'r') as ids_file:
            ids_lines = ids_file.readlines()

            for id_value in ids_lines:
                actual_id, username, *_ = id_value.split(' || ')
                if actual_id.strip() == cardID:
                    log("fetching username, CARD ID \"" + cardID + "\" = \"" + username.strip() + "\"")
                    return username.strip()  

            log("Failed to find username for \"" + cardID + "\"")
    except FileNotFoundError as e:
        log("ERROR: cant find Username file!")
        return f"Error: {e} - File not found."



def EvaluateRank(cardID): #hämtar vilken "rank" en person har 1 = deltagare (lägsta) 2 = Lärare 3 = Admin
    
    if (filecheck('./Data\\Deltagare.txt', cardID) == True):
        return 1
    elif (filecheck('./Data\\Larare.txt', cardID) == True):
        return 2
    elif (filecheck('./Data\\Admin.txt', cardID) == True):
        return 3
    else:
     return 0

def getPassword(cardID):  # hämtar lösenordet för att kolla om det stämmer
    rank = EvaluateRank(cardID)
    if rank == 0:
        return "Invalid Rank, please contact support"

    file = f'./Data/{["Deltagare", "Larare", "Admin"][rank-1]}.txt'
    passfile = file + ':passwords.txt'

    try:
        with open(file, 'r') as ids_file, open(passfile, 'r') as passwords_file:
            ids_lines = ids_file.readlines()
            passwords_lines = passwords_file.readlines()

            if len(ids_lines) != len(passwords_lines):
                log("warning! unequal passwords and usernames in " + passfile)
                raise ValueError("The number of IDs and passwords do not match")

            for id_value, password in zip(ids_lines, passwords_lines):
                actual_id = id_value.split(' || ')[0].strip()
                if actual_id == cardID:
                    
                    return password.strip()

            return "ERROR NO PASSWORD FOUND OLIVER OM DETTA HÄNDER STOPPA INLOGGET!!!"
    except FileNotFoundError as e:
        return f"Error: {e} - File not found."

def getEmail(cardID): #hämtar Mailen för att kolla om det stämmer(samma som lösenorder för inlogg)
    rank = EvaluateRank(cardID)
    if rank == 0:
        return "Invalid rank for the cardID"

    file = f'./Data/{["Deltagare", "Larare", "Admin"][rank-1]}.txt'
    
    try:
        with open(file, 'r') as ids_file:
            ids_lines = ids_file.readlines()

            for id_value in ids_lines:
                actual_id, name, email = map(str.strip, id_value.split(' || '))  
                if actual_id == cardID:
                    return email  

            return 0
    except FileNotFoundError as e:
        return 0



def narvaro(cardID):
    global datafile
    excel_path = datafile
    try:
        rank = EvaluateRank(cardID)
        if rank == 0:
            log("Invalid rank for the cardID")
            return 2

        username = fixSvenska(getUsername(cardID))  
        email = getEmail(cardID)       

        #excel_path = './Data/Narvaro/Narvaro_2024.xlsx'

        now = datetime.now()
        current_date = now.strftime("%Y-%m-%d")  
        week_number = now.isocalendar()[1]       
        current_weekday = now.strftime("%A")      
        current_time = now.strftime("%H:%M:%S")   

        day_str_swedish = {
            'Monday': 'Måndag',
            'Tuesday': 'Tisdag',
            'Wednesday': 'Onsdag',
            'Thursday': 'Torsdag',
            'Friday': 'Fredag',
            'Saturday': 'Lördag',
            'Sunday': 'Söndag'
        }[current_weekday]

        title = f"**Närvaro för {current_date} (Vecka {week_number}, {day_str_swedish})**"

        if not os.path.isfile(excel_path):
            workbook = openpyxl.Workbook()
            workbook.save(excel_path)

        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        date_header_written = False

        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == title:
                date_header_written = True
                break

        if not date_header_written:
            sheet.append([title])
            sheet.append(["Tid In (HH/MM/SS)", "Namn", "E-post", "Tid Ut (HH/MM/SS)"]) 

        sheet.append([current_time, username, email])

        

        try:
            workbook.save(excel_path)
            log(f"Närvaro registrerad för {username} kl. {current_time} den {current_date}.")
            return 0
        except PermissionError as e:
            log(f"Error: {e} cant open the file")
            return 3
        except Exception as e:
            log(f"An unexpected error occurred: {e}")
            return 3 

    except Exception as e:
        log(f"Error processing cardID {cardID}: {e}")




start_date = None
log_directory = './Data/Logs'

def log(message):
    global start_date

    if not os.path.exists(log_directory):
        os.makedirs(log_directory)

    log_file = os.path.join(log_directory, 'latestLog.txt')
    current_time = datetime.now()

    if not start_date:
        if os.path.isfile(log_file):
            old_start_date = datetime.fromtimestamp(os.path.getmtime(log_file))
            old_log_file_name = (
                f"log_{old_start_date.strftime('%Y%m%d')}_"
                f"{old_start_date.strftime('%H%M%S')}_to_"
                f"{current_time.strftime('%Y%m%d')}_"
                f"{current_time.strftime('%H%M%S')}.txt"
            )
            old_log_file_name = old_log_file_name.replace(":", "-")  
            os.rename(log_file, os.path.join(log_directory, old_log_file_name))

        start_date = current_time
        with open(log_file, 'w') as file:
            file.write(f"Logging started on: {start_date.strftime('%Y-%m-%d %H:%M:%S')}\n")
        print(f"Logging has started at {start_date.strftime('%Y-%m-%d %H:%M:%S')}.")

    timestamp = current_time.strftime('%Y-%m-%d %H:%M:%S')
    with open(log_file, 'a') as file:
        file.write(f"{timestamp} - {message}\n")
    print(f"Logged: {message} at {timestamp}.")

    if message.lower() == "stop logging":
        stop_date = datetime.now()
        with open(log_file, 'a') as file:
            file.write(f"Logging stopped on: {stop_date.strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        new_file_name = (
            f"log_{start_date.strftime('%Y%m%d')}_"
            f"{start_date.strftime('%H%M%S')}_to_"
            f"{stop_date.strftime('%Y%m%d')}_"
            f"{stop_date.strftime('%H%M%S')}.txt"
        )
        
        new_file_name = new_file_name.replace(":", "-")  
        os.rename(log_file, os.path.join(log_directory, new_file_name))
        
        print(f"Logging has stopped. Log saved as {new_file_name}.")
        start_date = None  


def EvaluateMail(email):
    
    if (filecheck('./Data\\Deltagare.txt', email) == True):
        return 1
    elif (filecheck('./Data\\Larare.txt', email) == True):
        return 2
    elif (filecheck('./Data\\Admin.txt', email) == True):
        return 3
    return 0

def checkpass(emailentered, passentered):  
    try:
        mailfile = EvaluateMail(emailentered)
        if mailfile == 0:
            log(f"Blocked suspicious password validation attempt! Email: {emailentered}")
            return False
        file = f'./Data/{["Deltagare", "Larare", "Admin"][mailfile - 1]}.txt'
    except Exception as e:
        log(f"Error evaluating email: {e}")
        return False

    try:
        with open(file, 'r') as ids_file:
            ids_lines = ids_file.readlines()

            for row, id_value in enumerate(ids_lines, start=1):
                actual_id, name, email = map(str.strip, id_value.split(' || '))
                if email == emailentered:
                    passfile = f'./Data/{["Deltagare.txt:Passwords", "Larare.txt:Passwords", "Admin.txt:Passwords"][mailfile - 1]}.txt'
                    try:
                        with open(passfile, 'r') as pass_file:
                            passwords = pass_file.readlines()

                            if row - 1 < len(passwords):  
                                passfound = passwords[row - 1].strip()  
                                if passfound == passentered:
                                    return True
                            return False
                    except FileNotFoundError as e:
                        log("Error opening password file: " + str(e))
                        return False  
            return False
    except FileNotFoundError as e:
        log("Error opening file: " + str(e))
        return False


def addMember(cardID, name, email, rank, password, authID):

    try:
        resultsofcardid = EvaluateRank(cardID)
        if resultsofcardid != 0:
            log(f"duplicate cardID atempted to make second user {resultsofcardid}")
            return 0
    except:
        log(f"Stoped attempt to create a duplicate user, cardID: {cardID}, email: {email}, name: {name}")

    try:
        ifname = getUsername(cardID)
        ifemail = getEmail(cardID)


        if(ifname == 0 or ifemail == 0):
            log("duplicate username or email!")
            return 0
    except:
        log(" no duplicate username or email!")

    
    rank_files = {
        1: './Data/Deltagare.txt',
        2: './Data/Larare.txt',
        3: './Data/Admin.txt'
    }

    if rank == 2 or rank == 3:
        authingRank = EvaluateRank(str(authID))
        log(f"authing rank returned: {authingRank}, Auth id is {authID}")
        if authingRank != 2 and authingRank != 3:  
            log(f"{authID} invalid auth to create high privilege user")
            return 0

        auther = getUsername(authID)
        
        if not auther:
            log(f"Failed to find username for {authID}")
            return 0
        
        log(f"{auther} authorized the addition of a high-privilege account.")

    file_path = rank_files.get(rank)
    if not file_path:
        log("Invalid rank provided.")
        return 0

    entry = f"{cardID} || {name} || {email}\n"

    password_filepath = f"{file_path}:Passwords.txt"
    try:
        with open(file_path, 'a') as file:
            file.write(entry)
        log(f"New user added \"{name}\" with rank {rank}")
        with open(password_filepath, 'a') as passwordFile:
            passwordFile.write(f"{password}\n")
        log(f"adding new password for {name}")
        return 1
    except Exception as e:
        log("Error adding new user!")
        return 0





def LookupNarvaro(cardID):
    global datafile
    file_path = datafile
    if EvaluateRank(cardID) == 0:
        log(f"Did not find user prompting new user cardID: {cardID}")
        return 2

    today_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")
    
    #file_path = './Data/Narvaro/Narvaro_2024.xlsx'
    
    if not os.path.exists(file_path):
        log("File does not exist.")
        return 3

    try:
        workbook = load_workbook(file_path)
    except PermissionError as e:
        log(f"Error: {e} cant open the file")
        return 4 

    attendance_sheet_name = 'Sheet1'
    
    if attendance_sheet_name not in workbook.sheetnames:
        log("Can't find Sheet1 in excel")
        return 3  
    
    sheet = workbook[attendance_sheet_name]

    email = getEmail(cardID)
    if email == "ERROR NO EMAIL FOUND FOR THIS ID!":
        log("No email found for cardID: ", cardID)
        return 3

    if email and "ERROR" not in email:
        found = False
        rows = list(sheet.iter_rows(min_row=2))  
        for row in reversed(rows):
            if row[2].value == email:
                if row[3].value:  
                    log("Tid Ut is already filled out for this user, adding user again...")
                    narvaro(cardID)
                    return 1
                row[3].value = current_time  
                found = True
                break

        if found:

                try:
                    workbook.save(file_path)
                    log("Avgång registrerat för " + getUsername(cardID))
                    return 0
                except PermissionError as e:
                    log(f"Error: {e} cant open the file")
                    return 4
                except Exception as e:
                    print(f"An unexpected error occurred: {e}")
                    return 3      
        else:
            narvaro(cardID)
            return 1 
    else:
        log("Invalid email from database")
        return 3